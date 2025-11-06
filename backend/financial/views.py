"""
Views do módulo Financeiro
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Sum, Q
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
from django.http import HttpResponse
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from core.permissions import IsSameTenant
from .models import PaymentMethod, Transaction, CashFlow
from .serializers import (
    PaymentMethodSerializer,
    TransactionSerializer,
    CreateTransactionSerializer,
    CashFlowSerializer
)


class PaymentMethodViewSet(viewsets.ModelViewSet):
    """ViewSet para gerenciamento de Métodos de Pagamento"""
    serializer_class = PaymentMethodSerializer
    permission_classes = [IsAuthenticated, IsSameTenant]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['is_active']

    def get_queryset(self):
        """Retorna apenas métodos de pagamento do mesmo tenant"""
        if self.request.user.is_authenticated:
            return PaymentMethod.objects.filter(tenant=self.request.user.tenant)
        return PaymentMethod.objects.none()

    @action(detail=False, methods=['get'])
    def active(self, request):
        """Retorna apenas métodos de pagamento ativos"""
        payment_methods = self.get_queryset().filter(is_active=True)
        serializer = self.get_serializer(payment_methods, many=True)
        return Response(serializer.data)


class TransactionViewSet(viewsets.ModelViewSet):
    """ViewSet para gerenciamento de Transações Financeiras"""
    serializer_class = TransactionSerializer
    permission_classes = [IsAuthenticated, IsSameTenant]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['type', 'payment_method', 'appointment']

    def get_queryset(self):
        """Retorna apenas transações do mesmo tenant com filtros opcionais - Otimizado"""
        if self.request.user.is_authenticated:
            queryset = Transaction.objects.filter(
                tenant=self.request.user.tenant
            ).select_related(
                'payment_method',
                'appointment__service',
                'appointment__customer',
                'created_by',
                'tenant'
            ).order_by('-date', '-created_at')
            
            # Filtro por data específica
            date = self.request.query_params.get('date', None)
            if date:
                queryset = queryset.filter(date=date)
            
            # Filtro por intervalo de datas
            start_date = self.request.query_params.get('start_date', None)
            end_date = self.request.query_params.get('end_date', None)
            if start_date and end_date:
                queryset = queryset.filter(
                    date__gte=start_date,
                    date__lte=end_date
                )
            
            return queryset
        return Transaction.objects.none()

    def create(self, request, *args, **kwargs):
        """Cria nova transação usando serializer simplificado"""
        serializer = CreateTransactionSerializer(
            data=request.data,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        transaction = serializer.save()

        # Retorna com o serializer completo
        output_serializer = TransactionSerializer(transaction)
        return Response(output_serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'])
    def today(self, request):
        """Retorna transações de hoje"""
        today = timezone.now().date()
        transactions = self.get_queryset().filter(date=today)
        serializer = self.get_serializer(transactions, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Retorna resumo financeiro do período"""
        queryset = self.get_queryset()
        
        # Aplica filtros de data se fornecidos
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if start_date and end_date:
            queryset = queryset.filter(date__gte=start_date, date__lte=end_date)
        else:
            # Padrão: últimos 30 dias
            today = timezone.now().date()
            start_date = today - timedelta(days=30)
            queryset = queryset.filter(date__gte=start_date)
        
        # Calcula totais
        revenue = queryset.filter(type='receita').aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')
        
        expenses = queryset.filter(type='despesa').aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')
        
        balance = revenue - expenses
        
        return Response({
            'start_date': start_date,
            'end_date': end_date,
            'total_revenue': revenue,
            'total_expenses': expenses,
            'balance': balance,
            'transaction_count': queryset.count()
        })

    @action(detail=False, methods=['get'])
    def by_payment_method(self, request):
        """Retorna resumo por método de pagamento"""
        queryset = self.get_queryset()
        
        # Aplica filtros de data
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if start_date and end_date:
            queryset = queryset.filter(date__gte=start_date, date__lte=end_date)
        
        # Agrupa por método de pagamento
        payment_methods = PaymentMethod.objects.filter(tenant=request.user.tenant)
        results = []
        
        for pm in payment_methods:
            pm_transactions = queryset.filter(payment_method=pm)
            revenue = pm_transactions.filter(type='receita').aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0.00')
            
            expenses = pm_transactions.filter(type='despesa').aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0.00')
            
            results.append({
                'payment_method': PaymentMethodSerializer(pm).data,
                'total_revenue': revenue,
                'total_expenses': expenses,
                'balance': revenue - expenses,
                'transaction_count': pm_transactions.count()
            })
        
        return Response(results)

    @action(detail=False, methods=['get'])
    def revenue_chart(self, request):
        """
        Retorna dados para gráfico de receita ao longo do tempo
        Query params: start_date, end_date, period (day|week|month)
        """
        from datetime import datetime, timedelta
        from django.db.models import Count
        from django.db.models.functions import TruncDate, TruncWeek, TruncMonth
        
        # Obter parâmetros
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        period = request.query_params.get('period', 'day')  # day, week, month
        
        # Padrão: últimos 30 dias
        if not start_date or not end_date:
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=30)
        
        # Filtrar transações de receita
        queryset = self.get_queryset().filter(
            type='receita',
            date__gte=start_date,
            date__lte=end_date
        )
        
        # Agrupar por período
        if period == 'month':
            data = queryset.annotate(
                period=TruncMonth('date')
            ).values('period').annotate(
                total=Sum('amount'),
                count=Count('id')
            ).order_by('period')
        elif period == 'week':
            data = queryset.annotate(
                period=TruncWeek('date')
            ).values('period').annotate(
                total=Sum('amount'),
                count=Count('id')
            ).order_by('period')
        else:  # day
            data = queryset.annotate(
                period=TruncDate('date')
            ).values('period').annotate(
                total=Sum('amount'),
                count=Count('id')
            ).order_by('period')
        
        return Response({
            'start_date': start_date,
            'end_date': end_date,
            'period': period,
            'data': list(data)
        })

    @action(detail=False, methods=['get'])
    def expense_chart(self, request):
        """
        Retorna dados para gráfico de despesas ao longo do tempo
        Query params: start_date, end_date, period (day|week|month)
        """
        from datetime import datetime, timedelta
        from django.db.models import Count
        from django.db.models.functions import TruncDate, TruncWeek, TruncMonth
        
        # Obter parâmetros
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        period = request.query_params.get('period', 'day')
        
        # Padrão: últimos 30 dias
        if not start_date or not end_date:
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=30)
        
        # Filtrar transações de despesa
        queryset = self.get_queryset().filter(
            type='despesa',
            date__gte=start_date,
            date__lte=end_date
        )
        
        # Agrupar por período
        if period == 'month':
            data = queryset.annotate(
                period=TruncMonth('date')
            ).values('period').annotate(
                total=Sum('amount'),
                count=Count('id')
            ).order_by('period')
        elif period == 'week':
            data = queryset.annotate(
                period=TruncWeek('date')
            ).values('period').annotate(
                total=Sum('amount'),
                count=Count('id')
            ).order_by('period')
        else:  # day
            data = queryset.annotate(
                period=TruncDate('date')
            ).values('period').annotate(
                total=Sum('amount'),
                count=Count('id')
            ).order_by('period')
        
        return Response({
            'start_date': start_date,
            'end_date': end_date,
            'period': period,
            'data': list(data)
        })

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Exporta transações em formato CSV"""
        queryset = self.filter_queryset(self.get_queryset())
        
        # Criar resposta CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="transacoes.csv"'
        response.write('\ufeff')  # BOM para UTF-8
        
        writer = csv.writer(response)
        writer.writerow([
            'Data',
            'Tipo',
            'Categoria',
            'Descrição',
            'Valor',
            'Método de Pagamento',
            'Agendamento'
        ])
        
        for transaction in queryset:
            writer.writerow([
                transaction.date.strftime('%d/%m/%Y'),
                transaction.get_type_display(),
                transaction.get_category_display(),
                transaction.description,
                f'R$ {transaction.amount:.2f}',
                transaction.payment_method.name if transaction.payment_method else '-',
                f'#{transaction.appointment.id}' if transaction.appointment else '-'
            ])
        
        return response

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Exporta transações em formato Excel"""
        queryset = self.filter_queryset(self.get_queryset())
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Transações"
        
        # Estilizar cabeçalho
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        
        headers = [
            'Data',
            'Tipo',
            'Categoria',
            'Descrição',
            'Valor',
            'Método de Pagamento',
            'Agendamento'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Adicionar dados
        for row, transaction in enumerate(queryset, start=2):
            ws.cell(row=row, column=1, value=transaction.date.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=transaction.get_type_display())
            ws.cell(row=row, column=3, value=transaction.get_category_display())
            ws.cell(row=row, column=4, value=transaction.description)
            ws.cell(row=row, column=5, value=f'R$ {transaction.amount:.2f}')
            ws.cell(row=row, column=6, value=transaction.payment_method.name if transaction.payment_method else '-')
            ws.cell(row=row, column=7, value=f'#{transaction.appointment.id}' if transaction.appointment else '-')
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 15
        
        # Criar resposta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="transacoes.xlsx"'
        wb.save(response)
        
        return response


class CashFlowViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet para visualização de Fluxo de Caixa"""
    serializer_class = CashFlowSerializer
    permission_classes = [IsAuthenticated, IsSameTenant]

    def get_queryset(self):
        """Retorna fluxo de caixa do tenant com filtros opcionais"""
        if self.request.user.is_authenticated:
            queryset = CashFlow.objects.filter(tenant=self.request.user.tenant)
            
            # Filtro por intervalo de datas
            start_date = self.request.query_params.get('start_date', None)
            end_date = self.request.query_params.get('end_date', None)
            if start_date and end_date:
                queryset = queryset.filter(
                    date__gte=start_date,
                    date__lte=end_date
                )
            
            return queryset
        return CashFlow.objects.none()

    @action(detail=False, methods=['post'])
    def calculate(self, request):
        """Calcula ou recalcula o fluxo de caixa para uma data"""
        date = request.data.get('date')
        if not date:
            return Response(
                {'error': 'Data é obrigatória'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Busca ou cria o registro de fluxo de caixa
        cash_flow, created = CashFlow.objects.get_or_create(
            tenant=request.user.tenant,
            date=date
        )
        
        # Calcula os valores
        cash_flow.calculate()
        
        serializer = self.get_serializer(cash_flow)
        return Response(serializer.data)
