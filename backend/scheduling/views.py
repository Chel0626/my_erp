"""
Views do módulo de Agendamentos
Implementa BLOCO 4: Workflows do módulo de Agendamentos
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from django.db.models import Q
from datetime import timedelta
from django.http import HttpResponse
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from .models import Service, Appointment
from .serializers import (
    ServiceSerializer,
    AppointmentSerializer,
    CreateAppointmentSerializer
)
from core.permissions import IsSameTenant, IsTenantAdmin


class ServiceViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gerenciamento de Serviços
    Implementa filtros automáticos por tenant
    """
    serializer_class = ServiceSerializer
    permission_classes = [IsAuthenticated, IsSameTenant]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['is_active']

    def get_queryset(self):
        """
        Retorna apenas serviços do mesmo tenant
        Implementa: BLOCO 3 - Regras de Segurança
        """
        if self.request.user.is_authenticated:
            return Service.objects.filter(tenant=self.request.user.tenant)
        return Service.objects.none()

    @action(detail=False, methods=['get'])
    def active(self, request):
        """Retorna apenas serviços ativos"""
        services = self.get_queryset().filter(is_active=True)
        serializer = self.get_serializer(services, many=True)
        return Response(serializer.data)


class AppointmentViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gerenciamento de Agendamentos
    Implementa: BLOCO 4 - Workflow: Criar Novo Agendamento
    """
    serializer_class = AppointmentSerializer
    permission_classes = [IsAuthenticated, IsSameTenant]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status', 'professional', 'service']

    def get_queryset(self):
        """
        Retorna apenas agendamentos do mesmo tenant
        Implementa: BLOCO 3 - Regras de Segurança
        """
        if self.request.user.is_authenticated:
            queryset = Appointment.objects.filter(tenant=self.request.user.tenant)
            
            # Filtros opcionais via query params
            date = self.request.query_params.get('date', None)
            if date:
                queryset = queryset.filter(start_time__date=date)
            
            start_date = self.request.query_params.get('start_date', None)
            end_date = self.request.query_params.get('end_date', None)
            if start_date and end_date:
                queryset = queryset.filter(
                    start_time__date__gte=start_date,
                    start_time__date__lte=end_date
                )
            
            return queryset
        return Appointment.objects.none()

    def create(self, request, *args, **kwargs):
        """
        Cria novo agendamento
        Usa serializer simplificado para melhor UX
        """
        serializer = CreateAppointmentSerializer(
            data=request.data,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        appointment = serializer.save()

        # Retorna com o serializer completo
        output_serializer = AppointmentSerializer(appointment)
        return Response(output_serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'])
    def today(self, request):
        """Retorna agendamentos do dia atual"""
        today = timezone.now().date()
        appointments = self.get_queryset().filter(start_time__date=today)
        serializer = self.get_serializer(appointments, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def week(self, request):
        """Retorna agendamentos da semana atual"""
        today = timezone.now().date()
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=6)
        
        appointments = self.get_queryset().filter(
            start_time__date__gte=week_start,
            start_time__date__lte=week_end
        )
        serializer = self.get_serializer(appointments, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def confirm(self, request, pk=None):
        """Confirma um agendamento"""
        appointment = self.get_object()
        appointment.status = 'confirmado'
        appointment.save()
        serializer = self.get_serializer(appointment)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def start(self, request, pk=None):
        """Inicia atendimento"""
        appointment = self.get_object()
        appointment.status = 'em_atendimento'
        appointment.save()
        serializer = self.get_serializer(appointment)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Conclui um agendamento"""
        appointment = self.get_object()
        appointment.status = 'concluido'
        appointment.save()
        serializer = self.get_serializer(appointment)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Cancela um agendamento"""
        appointment = self.get_object()
        appointment.status = 'cancelado'
        appointment.save()
        serializer = self.get_serializer(appointment)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def status_distribution(self, request):
        """
        Retorna distribuição de agendamentos por status
        Para gráfico de pizza/donut
        """
        from django.db.models import Count
        
        # Obter parâmetros de data
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        queryset = self.get_queryset()
        
        if start_date and end_date:
            queryset = queryset.filter(
                start_time__date__gte=start_date,
                start_time__date__lte=end_date
            )
        
        # Agrupar por status
        data = queryset.values('status').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Calcular total e percentuais
        total = sum(item['count'] for item in data)
        
        result = []
        for item in data:
            percentage = (item['count'] / total * 100) if total > 0 else 0
            result.append({
                'status': item['status'],
                'status_display': dict(Appointment.STATUS_CHOICES).get(item['status'], item['status']),
                'count': item['count'],
                'percentage': round(percentage, 2)
            })
        
        return Response({
            'total': total,
            'data': result
        })

    @action(detail=False, methods=['get'])
    def top_services(self, request):
        """
        Retorna top serviços mais agendados
        Query params: start_date, end_date, limit (padrão: 10)
        """
        from django.db.models import Count, Sum
        
        # Obter parâmetros
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        limit = int(request.query_params.get('limit', 10))
        
        queryset = self.get_queryset().filter(status='concluido')
        
        if start_date and end_date:
            queryset = queryset.filter(
                start_time__date__gte=start_date,
                start_time__date__lte=end_date
            )
        
        # Agrupar por serviço
        data = queryset.values(
            'service', 'service__name', 'service__price'
        ).annotate(
            count=Count('id'),
            total_revenue=Sum('service__price')
        ).order_by('-count')[:limit]
        
        result = []
        for item in data:
            result.append({
                'service_id': item['service'],
                'service_name': item['service__name'],
                'price': item['service__price'],
                'appointments_count': item['count'],
                'total_revenue': item['total_revenue']
            })
        
        return Response(result)

    @action(detail=False, methods=['get'])
    def professional_performance(self, request):
        """
        Retorna desempenho de cada profissional
        Query params: start_date, end_date
        """
        from django.db.models import Count, Sum, Avg
        from core.models import User
        
        # Obter parâmetros
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        queryset = self.get_queryset()
        
        if start_date and end_date:
            queryset = queryset.filter(
                start_time__date__gte=start_date,
                start_time__date__lte=end_date
            )
        
        # Buscar todos os profissionais do tenant
        professionals = User.objects.filter(tenant=request.user.tenant)
        
        result = []
        for professional in professionals:
            # Agendamentos do profissional
            prof_appointments = queryset.filter(professional=professional)
            
            total_appointments = prof_appointments.count()
            completed = prof_appointments.filter(status='concluido').count()
            cancelled = prof_appointments.filter(status='cancelado').count()
            
            # Calcular receita (apenas concluídos)
            revenue = prof_appointments.filter(status='concluido').aggregate(
                total=Sum('service__price')
            )['total'] or 0
            
            # Taxa de conclusão
            completion_rate = (completed / total_appointments * 100) if total_appointments > 0 else 0
            
            result.append({
                'professional_id': str(professional.id),
                'professional_name': professional.name,
                'professional_email': professional.email,
                'total_appointments': total_appointments,
                'completed': completed,
                'cancelled': cancelled,
                'total_revenue': float(revenue),
                'completion_rate': round(completion_rate, 2)
            })
        
        # Ordenar por receita total
        result.sort(key=lambda x: x['total_revenue'], reverse=True)
        
        return Response(result)

    @action(detail=False, methods=['get'])
    def appointments_timeline(self, request):
        """
        Retorna timeline de agendamentos para gráfico de linha
        Query params: start_date, end_date, period (day|week|month)
        """
        from django.db.models import Count
        from django.db.models.functions import TruncDate, TruncWeek, TruncMonth
        
        # Obter parâmetros
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        period = request.query_params.get('period', 'day')
        
        # Padrão: últimos 30 dias
        if not start_date or not end_date:
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=30)
        
        queryset = self.get_queryset().filter(
            start_time__date__gte=start_date,
            start_time__date__lte=end_date
        )
        
        # Agrupar por período
        if period == 'month':
            data = queryset.annotate(
                period=TruncMonth('start_time')
            ).values('period').annotate(
                total=Count('id'),
                confirmed=Count('id', filter=Q(status='confirmado')),
                completed=Count('id', filter=Q(status='concluido')),
                cancelled=Count('id', filter=Q(status='cancelado'))
            ).order_by('period')
        elif period == 'week':
            data = queryset.annotate(
                period=TruncWeek('start_time')
            ).values('period').annotate(
                total=Count('id'),
                confirmed=Count('id', filter=Q(status='confirmado')),
                completed=Count('id', filter=Q(status='concluido')),
                cancelled=Count('id', filter=Q(status='cancelado'))
            ).order_by('period')
        else:  # day
            data = queryset.annotate(
                period=TruncDate('start_time')
            ).values('period').annotate(
                total=Count('id'),
                confirmed=Count('id', filter=Q(status='confirmado')),
                completed=Count('id', filter=Q(status='concluido')),
                cancelled=Count('id', filter=Q(status='cancelado'))
            ).order_by('period')
        
        return Response({
            'start_date': start_date,
            'end_date': end_date,
            'period': period,
            'data': list(data)
        })

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Exporta agendamentos em formato CSV"""
        queryset = self.filter_queryset(self.get_queryset())
        
        # Criar resposta CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="agendamentos.csv"'
        response.write('\ufeff')  # BOM para UTF-8
        
        writer = csv.writer(response)
        writer.writerow([
            'Data/Hora',
            'Cliente',
            'Telefone',
            'Serviço',
            'Profissional',
            'Status',
            'Valor',
            'Observações'
        ])
        
        for appointment in queryset:
            writer.writerow([
                appointment.start_time.strftime('%d/%m/%Y %H:%M'),
                appointment.customer.name,
                appointment.customer.phone,
                appointment.service.name,
                appointment.professional.get_full_name(),
                appointment.get_status_display(),
                f'R$ {appointment.price:.2f}',
                appointment.notes or '-'
            ])
        
        return response

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Exporta agendamentos em formato Excel"""
        queryset = self.filter_queryset(self.get_queryset())
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Agendamentos"
        
        # Estilizar cabeçalho
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        
        headers = [
            'Data/Hora',
            'Cliente',
            'Telefone',
            'Serviço',
            'Profissional',
            'Status',
            'Valor',
            'Observações'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Adicionar dados
        for row, appointment in enumerate(queryset, start=2):
            ws.cell(row=row, column=1, value=appointment.start_time.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=2, value=appointment.customer.name)
            ws.cell(row=row, column=3, value=appointment.customer.phone)
            ws.cell(row=row, column=4, value=appointment.service.name)
            ws.cell(row=row, column=5, value=appointment.professional.get_full_name())
            ws.cell(row=row, column=6, value=appointment.get_status_display())
            ws.cell(row=row, column=7, value=f'R$ {appointment.price:.2f}')
            ws.cell(row=row, column=8, value=appointment.notes or '-')
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 40
        
        # Criar resposta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="agendamentos.xlsx"'
        wb.save(response)
        
        return response

