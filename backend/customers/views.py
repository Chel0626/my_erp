"""
Views para o módulo de Clientes
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from django.db.models import Count, Q, Sum, Avg, F
from datetime import datetime, timedelta
from django.http import HttpResponse
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from .models import Customer
from .serializers import (
    CustomerSerializer,
    CustomerListSerializer,
    CustomerStatsSerializer,
    CreateCustomerSerializer
)
from core.permissions import IsSameTenant


class CustomerViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gerenciar Clientes
    
    Endpoints:
    - GET /api/customers/ - Lista todos os clientes
    - POST /api/customers/ - Cria novo cliente
    - GET /api/customers/{id}/ - Detalhes de um cliente
    - PUT /api/customers/{id}/ - Atualiza cliente completo
    - PATCH /api/customers/{id}/ - Atualiza cliente parcial
    - DELETE /api/customers/{id}/ - Deleta cliente
    - GET /api/customers/{id}/stats/ - Estatísticas do cliente
    - GET /api/customers/{id}/appointments/ - Agendamentos do cliente
    - GET /api/customers/birthdays/ - Aniversariantes do mês
    - GET /api/customers/inactive/ - Clientes inativos
    - POST /api/customers/{id}/activate/ - Ativa cliente
    - POST /api/customers/{id}/deactivate/ - Desativa cliente
    """
    permission_classes = [IsAuthenticated, IsSameTenant]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    # Filtros
    filterset_fields = ['tag', 'is_active', 'gender']
    search_fields = ['name', 'phone', 'email', 'cpf']
    ordering_fields = ['name', 'created_at', 'last_visit']
    ordering = ['-created_at']
    
    def get_queryset(self):
        """Filtra clientes do tenant do usuário"""
        return Customer.objects.filter(
            tenant=self.request.user.tenant
        ).prefetch_related('appointments')
    
    def get_serializer_class(self):
        """Retorna serializer apropriado para cada ação"""
        if self.action == 'list':
            return CustomerListSerializer
        elif self.action == 'create':
            return CreateCustomerSerializer
        elif self.action == 'stats':
            return CustomerStatsSerializer
        return CustomerSerializer
    
    def perform_create(self, serializer):
        """Adiciona tenant automaticamente na criação"""
        serializer.save(tenant=self.request.user.tenant)
    
    @action(detail=True, methods=['get'])
    def stats(self, request, pk=None):
        """
        GET /api/customers/{id}/stats/
        Retorna estatísticas detalhadas do cliente
        """
        customer = self.get_object()
        
        # Busca appointments do cliente
        from scheduling.models import Appointment
        appointments = Appointment.objects.filter(
            tenant=request.user.tenant,
            customer_name=customer.name  # Assumindo que customer_name está no Appointment
        )
        
        # Calcula estatísticas
        total_appointments = appointments.count()
        completed = appointments.filter(status='concluido').count()
        cancelled = appointments.filter(status='cancelado').count()
        no_show = appointments.filter(status='falta').count()
        
        # Calcula gastos
        total_spent = 0.0
        service_count = {}
        professional_count = {}
        
        completed_apps = appointments.filter(status='concluido')
        for app in completed_apps:
            # Total gasto
            if hasattr(app, 'service_details') and app.service_details:
                price = app.service_details.get('price', 0)
                try:
                    total_spent += float(price)
                except (ValueError, TypeError):
                    pass
            
            # Serviço favorito
            service_name = app.service_details.get('name') if hasattr(app, 'service_details') and app.service_details else 'Desconhecido'
            service_count[service_name] = service_count.get(service_name, 0) + 1
            
            # Profissional favorito
            prof_name = app.professional_details.get('name') if hasattr(app, 'professional_details') and app.professional_details else 'Desconhecido'
            professional_count[prof_name] = professional_count.get(prof_name, 0) + 1
        
        # Favoritos
        favorite_service = max(service_count.items(), key=lambda x: x[1])[0] if service_count else None
        favorite_professional = max(professional_count.items(), key=lambda x: x[1])[0] if professional_count else None
        
        # Ticket médio
        average_ticket = total_spent / completed if completed > 0 else 0.0
        
        # Primeira e última visita
        first_visit = appointments.order_by('start_time').first()
        last_visit = appointments.order_by('-start_time').first()
        
        stats = {
            'total_appointments': total_appointments,
            'completed_appointments': completed,
            'cancelled_appointments': cancelled,
            'no_show_appointments': no_show,
            'total_spent': round(total_spent, 2),
            'average_ticket': round(average_ticket, 2),
            'first_visit': first_visit.start_time if first_visit else None,
            'last_visit': last_visit.start_time if last_visit else None,
            'favorite_service': favorite_service,
            'favorite_professional': favorite_professional,
        }
        
        serializer = CustomerStatsSerializer(stats)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def appointments(self, request, pk=None):
        """
        GET /api/customers/{id}/appointments/
        Lista todos os agendamentos do cliente
        """
        customer = self.get_object()
        
        from scheduling.models import Appointment
        from scheduling.serializers import AppointmentSerializer
        
        appointments = Appointment.objects.filter(
            tenant=request.user.tenant,
            customer_name=customer.name
        ).order_by('-start_time')
        
        serializer = AppointmentSerializer(appointments, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def birthdays(self, request):
        """
        GET /api/customers/birthdays/
        Lista clientes que fazem aniversário este mês
        """
        current_month = datetime.now().month
        
        customers = self.get_queryset().filter(
            birth_date__month=current_month,
            is_active=True
        ).order_by('birth_date__day')
        
        serializer = CustomerListSerializer(customers, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def inactive(self, request):
        """
        GET /api/customers/inactive/
        Lista clientes que não vêm há mais de 60 dias
        """
        sixty_days_ago = datetime.now() - timedelta(days=60)
        
        customers = self.get_queryset().filter(
            Q(last_visit__lt=sixty_days_ago) | Q(last_visit__isnull=True),
            is_active=True
        ).order_by('last_visit')
        
        serializer = CustomerListSerializer(customers, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def activate(self, request, pk=None):
        """
        POST /api/customers/{id}/activate/
        Ativa um cliente
        """
        customer = self.get_object()
        customer.is_active = True
        customer.save()
        
        serializer = self.get_serializer(customer)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def deactivate(self, request, pk=None):
        """
        POST /api/customers/{id}/deactivate/
        Desativa um cliente
        """
        customer = self.get_object()
        customer.is_active = False
        customer.save()
        
        serializer = self.get_serializer(customer)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """
        GET /api/customers/summary/
        Retorna resumo estatístico de clientes
        """
        queryset = self.get_queryset()
        
        summary = {
            'total': queryset.count(),
            'active': queryset.filter(is_active=True).count(),
            'inactive': queryset.filter(is_active=False).count(),
            'vip': queryset.filter(tag='VIP').count(),
            'regular': queryset.filter(tag='REGULAR').count(),
            'new': queryset.filter(tag='NOVO').count(),
            'birthdays_this_month': queryset.filter(
                birth_date__month=datetime.now().month,
                is_active=True
            ).count(),
        }
        
        return Response(summary)

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Exporta clientes em formato CSV"""
        queryset = self.filter_queryset(self.get_queryset())
        
        # Criar resposta CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="clientes.csv"'
        response.write('\ufeff')  # BOM para UTF-8
        
        writer = csv.writer(response)
        writer.writerow([
            'Nome',
            'Telefone',
            'E-mail',
            'CPF',
            'Data de Nascimento',
            'Gênero',
            'Tag',
            'Endereço',
            'Ativo',
            'Cadastrado em'
        ])
        
        for customer in queryset:
            writer.writerow([
                customer.name,
                customer.phone,
                customer.email or '-',
                customer.cpf or '-',
                customer.birth_date.strftime('%d/%m/%Y') if customer.birth_date else '-',
                customer.get_gender_display() if customer.gender else '-',
                customer.get_tag_display(),
                customer.address or '-',
                'Sim' if customer.is_active else 'Não',
                customer.created_at.strftime('%d/%m/%Y %H:%M')
            ])
        
        return response

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Exporta clientes em formato Excel"""
        queryset = self.filter_queryset(self.get_queryset())
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"
        
        # Estilizar cabeçalho
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        
        headers = [
            'Nome',
            'Telefone',
            'E-mail',
            'CPF',
            'Data de Nascimento',
            'Gênero',
            'Tag',
            'Endereço',
            'Ativo',
            'Cadastrado em'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Adicionar dados
        for row, customer in enumerate(queryset, start=2):
            ws.cell(row=row, column=1, value=customer.name)
            ws.cell(row=row, column=2, value=customer.phone)
            ws.cell(row=row, column=3, value=customer.email or '-')
            ws.cell(row=row, column=4, value=customer.cpf or '-')
            ws.cell(row=row, column=5, value=customer.birth_date.strftime('%d/%m/%Y') if customer.birth_date else '-')
            ws.cell(row=row, column=6, value=customer.get_gender_display() if customer.gender else '-')
            ws.cell(row=row, column=7, value=customer.get_tag_display())
            ws.cell(row=row, column=8, value=customer.address or '-')
            ws.cell(row=row, column=9, value='Sim' if customer.is_active else 'Não')
            ws.cell(row=row, column=10, value=customer.created_at.strftime('%d/%m/%Y %H:%M'))
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 40
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 18
        
        # Criar resposta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="clientes.xlsx"'
        wb.save(response)
        
        return response

